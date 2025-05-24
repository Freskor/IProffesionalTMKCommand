import re
from excel_handler import (       
    write_cell, get_table_data_and_metadata,
    set_column_header, find_first_empty_cell_in_column,
    load_workbook_and_sheet, save_workbook,
    push_undo_state, pop_undo_state, apply_data_to_sheet
)
from openpyxl.utils import get_column_letter, column_index_from_string

def parse_cell_address(text):
    match = re.search(r"([a-zA-Z]+)(\d+)", text)
    if match:
        col_str = match.group(1).upper()
        row_str = match.group(2)
        try:
            return {"col": column_index_from_string(col_str), "row": int(row_str)}
        except ValueError:
            return None
    return None

def parse_and_execute(filename, command_text, active_cell, table_metadata):
    command_text = command_text.lower().strip()
    words = command_text.split()
    
    current_active_cell = active_cell.copy()         
    new_active_cell = active_cell.copy()
    refresh_data = False
    message = "Команда не распознана."
    success = False
    
    current_data_for_undo = get_table_data_and_metadata(filename)['data']


    max_rows = table_metadata.get('max_row', current_active_cell['row'] + 5)           
    max_cols = table_metadata.get('max_col', current_active_cell['col'] + 5)

    if command_text == "вверх":
        if new_active_cell['row'] > 1:
            new_active_cell['row'] -= 1
        success = True
    elif command_text == "Вниз.":
        if new_active_cell['row'] < max_rows:               
             new_active_cell['row'] += 1
        else:                         
            new_active_cell['row'] +=1                 
        success = True
    elif command_text == "влево":
        if new_active_cell['col'] > 1:
            new_active_cell['col'] -= 1
        success = True
    elif command_text == "вправо":
        if new_active_cell['col'] < max_cols:
            new_active_cell['col'] += 1
        else:
            new_active_cell['col'] += 1                 
        success = True

    elif command_text in ["назад", "отмена"]:
        previous_state = pop_undo_state(filename)
        if previous_state:
            if apply_data_to_sheet(filename, previous_state):
                success = True
                message = "Последнее действие отменено."
                refresh_data = True
            else:
                message = "Не удалось отменить действие."
        else:
            message = "Нет действий для отмены."
    elif (words[0] == "столбец" and len(words) >= 3):
        col_ref = words[1]
        name_to_set = " ".join(words[2:])
        col_idx_to_set = -1
        try:
            if re.match(r"^[a-zA-Z]+$", col_ref):       
                col_idx_to_set = column_index_from_string(col_ref.upper())
            elif re.match(r"^\d+$", col_ref):       
                col_idx_to_set = int(col_ref)
            
            if col_idx_to_set > 0:
                push_undo_state(filename, current_data_for_undo)
                if set_column_header(filename, col_idx_to_set, name_to_set):
                    success = True
                    message = f"Столбцу {col_ref.upper()} присвоено имя '{name_to_set}'."
                    refresh_data = True
                else:
                    message = f"Не удалось установить имя столбца {col_ref.upper()}."
            else:
                message = f"Некорректный идентификатор столбца: {col_ref}"
        except ValueError:
            message = f"Некорректный идентификатор столбца: {col_ref}"

    elif len(words) == 1 and not re.match(r"(вверх|вниз|влево|вправо|назад|отмена)", words[0]):
        potential_col_name = words[0]
        found_cell = find_first_empty_cell_in_column(filename, potential_col_name)
        if found_cell:
            new_active_cell = found_cell
            success = True
            message = f"Переход к первой пустой ячейке в столбце '{potential_col_name}'."
    elif (len(words) == 1 and parse_cell_address(words[0])) or \
         (len(words) == 2 and words[0] == "ячейка" and parse_cell_address(words[1])):
        cell_str = words[-1]
        parsed_addr = parse_cell_address(cell_str)
        if parsed_addr:
            new_active_cell = parsed_addr
            success = True
            message = f"Активная ячейка: {cell_str.upper()}"
        else:
            message = f"Неверный адрес ячейки: {cell_str}"

    elif words[0] == "записать" and len(words) >= 2:
        value_to_write = " ".join(words[1:])
        target_cell_coords = current_active_cell           
        
        match_explicit_cell = re.search(r"\s+в\s+([a-zA-Z]+\d+)$", value_to_write)
        if match_explicit_cell:
            explicit_cell_str = match_explicit_cell.group(1)
            parsed_explicit_addr = parse_cell_address(explicit_cell_str)
            if parsed_explicit_addr:
                target_cell_coords = parsed_explicit_addr
                value_to_write = value_to_write[:match_explicit_cell.start()].strip()
        
        push_undo_state(filename, current_data_for_undo)
        if write_cell(filename, target_cell_coords['row'], target_cell_coords['col'], value_to_write):
            success = True
            message = f"Записано '{value_to_write}' в ячейку {get_column_letter(target_cell_coords['col'])}{target_cell_coords['row']}"
            refresh_data = True
            new_active_cell = target_cell_coords                   
        else:
            message = f"Не удалось записать в ячейку."


    elif words[0] == "поиск" and len(words) >= 2:
        search_term = " ".join(words[1:])
        wb, sheet, _ = load_workbook_and_sheet(filename)
        found = False
        if sheet:
            for r in range(1, sheet.max_row + 1):
                for c in range(1, sheet.max_column + 1):
                    cell_val = sheet.cell(row=r, column=c).value
                    if cell_val and search_term in str(cell_val).lower():
                        new_active_cell = {"row": r, "col": c}
                        message = f"Найдено '{search_term}' в ячейке {get_column_letter(c)}{r}."
                        success = True
                        found = True
                        break
                if found: break
            if not found:
                message = f"'{search_term}' не найдено."
        else:
            message = "Не удалось открыть таблицу для поиска."


    elif words[0] == "рассчитать" and words[1] == "сумму" and words[2] == "столбец" and len(words) >= 4:
        col_ref_sum = words[3]
        wb, sheet, _ = load_workbook_and_sheet(filename)
        if not sheet:
            message = "Не удалось открыть таблицу для расчета."
        else:
            target_col_idx_sum = -1
            if re.match(r"^[a-zA-Z]+$", col_ref_sum):
                try: target_col_idx_sum = column_index_from_string(col_ref_sum.upper())
                except: pass
            elif re.match(r"^\d+$", col_ref_sum):
                target_col_idx_sum = int(col_ref_sum)
            else:       
                for c_idx in range(1, sheet.max_column + 1):
                    header = sheet.cell(row=1, column=c_idx).value
                    if header and header.lower() == col_ref_sum.lower():
                        target_col_idx_sum = c_idx
                        break
            
            if target_col_idx_sum > 0:
                last_data_row = 0
                for r_idx in range(sheet.max_row, 0, -1):         
                    cell_val = sheet.cell(row=r_idx, column=target_col_idx_sum).value
                    if isinstance(cell_val, (int, float)):
                        last_data_row = r_idx
                        break
                if last_data_row == 0:           
                    last_data_row = 1                 
                
                formula_row = last_data_row + 1
                start_row_for_sum = 2 if sheet.cell(1, target_col_idx_sum).value else 1
                
                push_undo_state(filename, current_data_for_undo)

                formula = f"=SUM({get_column_letter(target_col_idx_sum)}{start_row_for_sum}:{get_column_letter(target_col_idx_sum)}{last_data_row})"
                if write_cell(filename, formula_row, target_col_idx_sum, formula, sheet_instance=sheet, workbook_instance=wb):
                    save_workbook(filename, wb)           
                    success = True
                    message = f"Формула '{formula}' вставлена в {get_column_letter(target_col_idx_sum)}{formula_row}."
                    refresh_data = True
                    new_active_cell = {"row": formula_row, "col": target_col_idx_sum}
                else:
                    message = "Не удалось вставить формулу."
            else:
                message = f"Не найден столбец '{col_ref_sum}' для расчета суммы."

    if success and command_text not in ["назад", "отмена"]:             
         pass


    return {
        "success": success,
        "message": message,
        "new_active_cell": new_active_cell,
        "refresh_data": refresh_data
    }