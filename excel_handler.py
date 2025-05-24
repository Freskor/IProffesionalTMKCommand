import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import os
import uuid
from datetime import datetime
import logging

TABLES_DIR = os.path.join("static", "tables")
os.makedirs(TABLES_DIR, exist_ok=True)

LAST_MODIFIED_TIMES = {}

logger = logging.getLogger(__name__)         
logger.setLevel(logging.DEBUG)
DEFAULT_ROWS = 100 
DEFAULT_COLS = 26   

def convert_sheet_to_luckysheet_celldata(sheet):
    celldata = []
    for r_idx, row in enumerate(sheet.iter_rows()):
        for c_idx, cell in enumerate(row):
            if cell.value is not None:
                value_str = str(cell.value)
                is_formula = value_str.startswith('=')
                
                cell_obj = {
                    "r": r_idx,
                    "c": c_idx,
                    "v": {
                        "v": cell.value,           
                        "m": value_str         
                    }
                }
                if is_formula:
                    cell_obj["v"]["f"] = value_str     
                
                celldata.append(cell_obj)
    return celldata

def get_luckysheet_data_from_excel(filename):
    workbook, sheet, _ = load_workbook_and_sheet(filename)
    if not sheet:
        return [{
            "name": "Sheet1",
            "celldata": [],
            "order": 0,
            "index": 0,
            "status": 1,
        }]

    celldata = convert_sheet_to_luckysheet_celldata(sheet)
    
    luckysheet_file_data = [
        {
            "name": "Cell",
            "celldata": celldata,
            "order": 0,       
            "index": "0",         
            "status": 1,             
            "config": {},               
            "zoomRatio": 1,
        }
    ]
    return luckysheet_file_data

def save_luckysheet_data_to_excel(filename, luckysheet_data_array):
    """
    Сохраняет данные из формата Luckysheet (массив объектов листов) в файл Excel.
    Возвращает кортеж (bool: успех, str: сообщение).
    """
    path = get_table_path(filename)
    logger.info(f"Начало сохранения данных Luckysheet в файл: {path}")

    if not luckysheet_data_array or not isinstance(luckysheet_data_array, list):
        msg = "Нет данных для сохранения или неверный формат (ожидался список листов)."
        logger.error(msg + f" Получено: {type(luckysheet_data_array)}")
        return False, msg

    try:
        if os.path.exists(path):
            workbook = openpyxl.load_workbook(path)
            logger.debug(f"Загружен существующий workbook: {filename}")
            for sheet_name in workbook.sheetnames:
                workbook.remove(workbook[sheet_name])
            logger.debug(f"Все существующие листы в {filename} удалены перед записью новых.")
        else:
            workbook = openpyxl.Workbook()
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])
            logger.debug(f"Создан новый workbook для: {filename}")

        if not luckysheet_data_array:                   
            sheet_name = "Sheet1"
            workbook.create_sheet(title=sheet_name)
            logger.info(f"Данные Luckysheet не содержат листов. Создан пустой лист '{sheet_name}' в {filename}.")

        for sheet_data_obj in luckysheet_data_array:
            if not isinstance(sheet_data_obj, dict):
                logger.warning(f"Пропущен элемент в luckysheet_data_array, так как это не словарь: {sheet_data_obj}")
                continue

            sheet_name = sheet_data_obj.get("name", f"Sheet_{sheet_data_obj.get('index', 'unknown')}")
            celldata = sheet_data_obj.get("celldata", [])
            sheet = workbook.create_sheet(title=sheet_name)
            logger.debug(f"Создан лист '{sheet_name}' в workbook.")

            if not celldata:
                logger.info(f"Лист '{sheet_name}' не содержит celldata (пустой лист).")
                continue

            max_r_written = -1
            max_c_written = -1

            for cell_obj in celldata:
                if not isinstance(cell_obj, dict):
                    logger.warning(f"Пропущен элемент в celldata листа '{sheet_name}', так как это не словарь: {cell_obj}")
                    continue

                r = cell_obj.get("r")        
                c = cell_obj.get("c")        
                v_data = cell_obj.get("v")     

                if r is None or c is None or v_data is None:
                    logger.warning(f"Пропущена ячейка в листе '{sheet_name}' из-за отсутствия r, c или v: {cell_obj}")
                    continue
                
                excel_row, excel_col = r + 1, c + 1

                value_to_write = None
                if isinstance(v_data, dict):
                    if "f" in v_data and v_data["f"] is not None:        
                        value_to_write = str(v_data["f"])         
                        if not value_to_write.startswith('='):
                            value_to_write = '=' + value_to_write         
                    elif "v" in v_data:         
                        value_to_write = v_data["v"]
                else:                     
                    value_to_write = v_data


                if value_to_write is not None:
                    try:
                        sheet.cell(row=excel_row, column=excel_col, value=value_to_write)
                        max_r_written = max(max_r_written, excel_row)
                        max_c_written = max(max_c_written, excel_col)
                    except Exception as cell_ex:
                        logger.error(f"Ошибка записи в ячейку ({excel_row},{excel_col}) листа '{sheet_name}': {cell_ex}. Значение: {value_to_write}")
            
            logger.info(f"Завершена обработка celldata для листа '{sheet_name}'. Макс. записанная ячейка: ({max_r_written}, {max_c_written})")

        workbook.save(path)
        LAST_MODIFIED_TIMES[filename] = datetime.now()
        msg = f"Таблица {filename} успешно сохранена (версия Luckysheet)."
        logger.info(msg + f" Время: {LAST_MODIFIED_TIMES[filename]}")
        return True, msg

    except Exception as e:
        msg = f"Ошибка при сохранении файла {filename} из данных Luckysheet: {e}"
        logger.error(msg, exc_info=True)
        return False, msg

def get_luckysheet_data_from_excel(filename):
    path = get_table_path(filename)
    logger.info(f"Загрузка данных из Excel ({filename}) для формата Luckysheet.")
    
    luckysheet_file_data = []

    if not os.path.exists(path):
        logger.warning(f"Файл {filename} не найден. Возвращаем структуру для пустого листа Luckysheet.")
        luckysheet_file_data.append({
            "name": "Sheet1",       
            "celldata": [],
            "order": 0,
            "index": "0",         
            "status": 1,      
            "config": {},
            "zoomRatio": 1,
        })
        return luckysheet_file_data

    try:
        workbook = openpyxl.load_workbook(path, data_only=False)         
        
        for idx, sheet_name in enumerate(workbook.sheetnames):
            sheet = workbook[sheet_name]
            celldata = []
            
            min_row, min_col, max_row, max_col = (None, None, None, None)
            if sheet.calculate_dimension() != 'A1:A1' or sheet['A1'].value is not None :             
                dims = sheet.calculate_dimension()   
                if dims:
                    try:
                        min_col_str, min_row_str, max_col_str, max_row_str = openpyxl.utils.range_boundaries(dims)
                        min_row, min_col, max_row, max_col = int(min_row_str), int(min_col_str), int(max_row_str), int(max_col_str)
                    except:               
                         min_row, min_col, max_row, max_col = 1, 1, sheet.max_row, sheet.max_column
            else:     
                min_row, min_col, max_row, max_col = 1, 1, 0, 0         


            for r_idx_excel in range(min_row, max_row + 1):     
                for c_idx_excel in range(min_col, max_col + 1):     
                    cell = sheet.cell(row=r_idx_excel, column=c_idx_excel)
                    if cell.value is not None:
                        r_luckysheet, c_luckysheet = r_idx_excel - 1, c_idx_excel - 1
                        
                        value_obj = {"v": cell.value, "m": str(cell.value)}         
                        
                        if cell.data_type == 'f':           
                            value_obj["f"] = str(cell.value)       
                            value_obj["v"] = str(cell.value)             
                        elif cell.is_date:
                            pass 
                        cell_luckysheet_obj = {"r": r_luckysheet, "c": c_luckysheet, "v": value_obj}
                        celldata.append(cell_luckysheet_obj)
            
            sheet_luckysheet_obj = {
                "name": sheet.title,
                "celldata": celldata,
                "order": idx,
                "index": str(idx),           
                "status": 1 if idx == 0 else 0,                
                "config": {},                 
                "zoomRatio": 1,
            }
            luckysheet_file_data.append(sheet_luckysheet_obj)
        
        if not luckysheet_file_data:               
            logger.warning(f"Книга {filename} не содержала листов. Возвращаем пустой лист.")
            luckysheet_file_data.append({ "name": "Sheet1", "celldata": [], "order": 0, "index": "0", "status": 1, "config":{}, "zoomRatio":1})

        logger.info(f"Успешно загружено {len(luckysheet_file_data)} листов из {filename} для Luckysheet.")
        return luckysheet_file_data

    except Exception as e:
        logger.error(f"Ошибка при загрузке файла {filename} для Luckysheet: {e}", exc_info=True)
        return [{ "name": "ErrorSheet", "celldata": [{"r":0, "c":0, "v":{"m":"Ошибка загрузки файла"}}], "order": 0, "index": "0", "status": 1, "config":{}, "zoomRatio":1}]
    
    print(f"Данные первого листа для сохранения: {luckysheet_data_array[0].get('name')}, количество ячеек: {len(luckysheet_data_array[0].get('celldata', []))}")
    sheet_data_obj = luckysheet_data_array[0]
    celldata = sheet_data_obj.get("celldata", [])

    workbook, old_sheet, path = load_workbook_and_sheet(filename)         

    if not workbook:             
        workbook = openpyxl.Workbook()
        if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
            std_sheet = workbook["Sheet"]
            workbook.remove(std_sheet)
            sheet = workbook.create_sheet(title=sheet_data_obj.get("name", "Sheet1"), index=0)       
        path = get_table_path(filename)
    else:
        sheet_name_to_save = sheet_data_obj.get("name", "Sheet1")
        if sheet_name_to_save in workbook.sheetnames:
            sheet = workbook[sheet_name_to_save]
            for row_obj in sheet.iter_rows():
                for cell_obj in row_obj:
                    cell_obj.value = None
        else:
            sheet = workbook.create_sheet(title=sheet_name_to_save, index=0)       

    sheet.title = sheet_data_obj.get("name", sheet.title)

    for cell_obj in celldata:
        r = cell_obj.get("r") 
        c = cell_obj.get("c")
        v_data = cell_obj.get("v")

        if r is not None and c is not None and v_data is not None:
            value_to_write = None
            if "f" in v_data and v_data["f"]:       
                 value_to_write = v_data["f"]
            elif "v" in v_data:
                 value_to_write = v_data["v"]
            
            if value_to_write is not None:
                sheet.cell(row=r + 1, column=c + 1, value=value_to_write)
    try:
        workbook.save(path)
        LAST_MODIFIED_TIMES[filename] = datetime.now()
        return True
    except Exception as e:
        print(f"Error saving workbook {filename} from Luckysheet data: {e}")
        return False
    LAST_MODIFIED_TIMES
def get_table_path(filename):
    return os.path.join(TABLES_DIR, filename)

def create_new_table(rows, cols):                   
    unique_id = uuid.uuid4()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"table_{timestamp}_{unique_id}.xlsx"
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1" 
    
    try:
        workbook.save(get_table_path(filename))
        LAST_MODIFIED_TIMES[filename] = datetime.now()
        return filename, {"row": 1, "col": 1, "rows_count": int(rows), "cols_count": int(cols)}
    except Exception as e:
        print(f"Error saving new table {filename}: {e}")
        return None, None
def save_full_table_data(filename, data_array):
    """
    Полностью перезаписывает активный лист предоставленным массивом данных.
    data_array - это массив массивов (строки ячеек).
    """
    workbook, sheet, path = load_workbook_and_sheet(filename)
    if not sheet:
        print(f"Sheet not found for {filename} in save_full_table_data")
        return False

    if sheet.max_row > 0 :
        for row in sheet.iter_rows():
            for cell in row:
                cell.value = None
    
    for r_idx, row_data in enumerate(data_array):
        if row_data is None: continue                 
        for c_idx, cell_value in enumerate(row_data):
            try:
                sheet.cell(row=r_idx + 1, column=c_idx + 1, value=cell_value)
            except Exception as cell_e:
                print(f"Error writing cell ({r_idx+1},{c_idx+1}) with value '{cell_value}': {cell_e}")
    try:
        workbook.save(path)
        LAST_MODIFIED_TIMES[filename] = datetime.now()       
        print(f"Full table data for {filename} saved successfully at {datetime.now()}")
        return True
    except Exception as e:
        print(f"Error saving workbook {filename} in save_full_table_data: {e}")
        return False

def load_workbook_and_sheet(filename):
    path = get_table_path(filename)
    if not os.path.exists(path):
        return None, None, None
    try:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        return workbook, sheet, path
    except Exception as e:
        print(f"Error loading workbook {filename}: {e}")
        return None, None, None

def get_table_data_and_metadata(filename):
    workbook, sheet, _ = load_workbook_and_sheet(filename)
    if not sheet:
        return None

    data = []
    max_r = sheet.max_row
    max_c = sheet.max_column

    if max_r == 1 and max_c == 1 and sheet.cell(1,1).value is None:
        pass                   

    for r in range(1, max_r + 1):
        row_data = []
        for c in range(1, max_c + 1):
            cell_value = sheet.cell(row=r, column=c).value
            row_data.append(cell_value if cell_value is not None else "")           
        data.append(row_data)
    
    if not data:
        data.append([""])
        
    return {
        "data": data,
        "filename": filename,
        "max_row": max_r,       
        "max_col": max_c        
    }

def write_cell(filename, row, col, value, sheet_instance=None, workbook_instance=None):
    wb_provided = workbook_instance is not None and sheet_instance is not None
    if not wb_provided:
        workbook, sheet, path = load_workbook_and_sheet(filename)
        if not sheet: return False
    else:
        sheet = sheet_instance
        workbook = workbook_instance
        path = get_table_path(filename)

    try:
        sheet.cell(row=int(row), column=int(col), value=value)
        if not wb_provided:                 
            workbook.save(path)
        LAST_MODIFIED_TIMES[filename] = datetime.now()
        return True
    except Exception as e:
        print(f"Error writing to cell in {filename} ({row},{col}): {e}")
        return False

def save_workbook(filename, workbook):
    try:
        workbook.save(get_table_path(filename))
        LAST_MODIFIED_TIMES[filename] = datetime.now()
        print(f"Workbook {filename} saved at {datetime.now()}")
        return True
    except Exception as e:
        print(f"Error saving workbook {filename}: {e}")
        return False

def get_column_header(filename, col_index):
    _, sheet, _ = load_workbook_and_sheet(filename)
    if not sheet or col_index < 1 or col_index > sheet.max_column:
        return None
    return sheet.cell(row=1, column=col_index).value

def set_column_header(filename, col_index, header_name, sheet_instance=None, workbook_instance=None):
    return write_cell(filename, 1, col_index, header_name, sheet_instance, workbook_instance)

def find_first_empty_cell_in_column(filename, column_name_or_index):
    workbook, sheet, _ = load_workbook_and_sheet(filename)
    if not sheet: return None

    target_col_idx = -1
    if isinstance(column_name_or_index, str):
        for c_idx in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=c_idx).value
            if header and header.lower() == column_name_or_index.lower():
                target_col_idx = c_idx
                break
        if target_col_idx == -1: return None       
    elif isinstance(column_name_or_index, int):
        target_col_idx = column_name_or_index
        if not (1 <= target_col_idx <= sheet.max_column): return None
    else:
        return None

    for r_idx in range(1, sheet.max_row + 2):                   
        cell_value = sheet.cell(row=r_idx, column=target_col_idx).value
        if cell_value is None or str(cell_value).strip() == "":
            return {"row": r_idx, "col": target_col_idx}
    return None                 

def list_excel_files():
    files = [f for f in os.listdir(TABLES_DIR) if f.endswith('.xlsx') and os.path.isfile(get_table_path(f))]
    return sorted(files, key=lambda f: os.path.getmtime(get_table_path(f)), reverse=True)

UNDO_STACK = {}     

def push_undo_state(filename, data):
    if filename not in UNDO_STACK:
        UNDO_STACK[filename] = []
    if len(UNDO_STACK[filename]) > 5:
        UNDO_STACK[filename].pop(0)
    UNDO_STACK[filename].append(data)

def pop_undo_state(filename):
    if filename in UNDO_STACK and UNDO_STACK[filename]:
        return UNDO_STACK[filename].pop()
    return None

def apply_data_to_sheet(filename, data_state):
    workbook, sheet, path = load_workbook_and_sheet(filename)
    if not sheet: return False
    
    for row_obj in sheet.iter_rows():
        for cell_obj in row_obj:
            cell_obj.value = None

    for r_idx, row_data in enumerate(data_state):
        for c_idx, cell_value in enumerate(row_data):
            sheet.cell(row=r_idx + 1, column=c_idx + 1, value=cell_value)
    
    workbook.save(path)
    LAST_MODIFIED_TIMES[filename] = datetime.now()
    return True